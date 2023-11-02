import csv
#from xlsxwriter.workbook import Workbook
import argparse
import pandas as pd
import os
from openpyxl import load_workbook

def add_worksheet(output,input):
	if os.path.isfile(output) == True:
		book = load_workbook(output)
		writer = pd.ExcelWriter(output, engine='openpyxl')
		writer.book = book
	else:
		writer = pd.ExcelWriter(output, engine='openpyxl')
	df = pd.read_csv(input, sep='\t')
	df.to_excel(writer, sheet_name=os.path.basename(input).split('.')[2], index=False)
	writer.save()
	
def test(input,output):
	# Create an XlsxWriter workbook object and add a worksheet
	workbook = Workbook(output)
	worksheet = workbook.add_worksheet()

	# Create a TSV file reader.
	tsv_reader = csv.reader(open(input,'rt'),delimiter="\t")

	# Read the row data from the TSV file and write it to the XLSX file
	for row, data in enumerate(tsv_reader):
		worksheet.write_row(row, 0, data)

	# Close the XLSX file
	workbook.close()

if __name__ == '__main__':
	parser = argparse.ArgumentParser()
	parser.add_argument('--out', required=True, help="xlsx file with the result")
	parser.add_argument('--input', required=True, help='All summary files')

	args = parser.parse_args()
	add_worksheet(args.out,args.input)
